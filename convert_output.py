import re
import json
import openpyxl
import ast
import string
from xlsxwriter import Workbook
from typing import List, Dict, Any, Tuple

INPUT_FILE_PATH_HAN_NOM_CLC = "data/HDTGDS_clean.hannom.clc.json"
INPUT_FILE_PATH_HAN_NOM_KANDIANGUJI = "data/HDTGDS_clean.hannom.json"
INPUT_FILE_PATH_QUOC_NGU = "data/HDTGDS_clean.quocngu.json"
OUTPUT_FILE_PATH = "result/result.xlsx"
QUOCNGU_SINONOM_DIC_FILE = "data/dic/QuocNgu_SinoNom_Dic.xlsx"
SINONOM_SIMILAR_DIC_FILE = "data/dic/SinoNom_similar_Dic.xlsx"

COL_IMAGE_NAME = 0
COL_ID = 1
COL_IMAGE_BOX = 2
COL_SINONOM_OCR = 3
COL_SINONOM_CHAR = 4
COL_CHU_QUOC_NGU = 5

# Đọc dữ liệu hán nôm ocr từ file JSON
def read_input_file_han_nom(file_path):
    with open(file_path, "r", encoding="utf-8") as file:
        data = json.load(file)
        return data
    
def read_input_file_quoc_ngu(file_path):
    with open(file_path, "r", encoding="utf-8") as file:
        data = json.load(file)
        return data

# Tạo workbook để ghi vào file xlsx
def create_workbook(filename):
    workbook = Workbook(filename)
    worksheet = workbook.add_worksheet("SinoNom Data")
    return workbook, worksheet

# Khởi tạo header của từng cột
def set_headers():
    headers = ["Image_name", "ID", "Image Box", "SinoNom OCR", "SinoNom Char", "Chữ Quốc Ngữ"]
    bold_format = workbook.add_format({'bold': True})
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, bold_format)

# Tải dữ liệu từ điển (SinoNom_similar_Dic & QuocNgu_SinoNom_Dic) vào  list 
def load_dictionary(file_path, key1, key2):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    return [
        {key1: row[0], key2: row[1]}
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]

# Xử lý định dạng chữ quốc ngữ 
def preprocess_quocngu_text(text):
    text = text.lower()
    text = text.translate(str.maketrans('', '', string.punctuation))
    return text.split()


cache_dp = None

# Tính ma trận Levenshtein
def compute_levenshtein_matrix(str1, str2):
    m, n = len(str1), len(str2)
    global cache_dp

    # Nếu cache_dp chưa được khởi tạo
    if not cache_dp:
        dp = [[0] * (n + 1) for _ in range(m + 1)]

        for i in range(m + 1):
            dp[i][0] = i
        for j in range(n + 1):
            dp[0][j] = j
        
        for i in range(1, m + 1):
            for j in range(1, n + 1):
                # val=2 nếu kết quả giao bằng 0
                # val=0 nết kết quả giao >= 1
                val = 2 if intersect_characters(str1[i-1], str2[j-1]) == 0 else 0
                dp[i][j] = min( dp[i-1][j]+1,           # deletion
                                dp[i][j-1]+1,           # insertion
                                dp[i-1][j-1] + val)     # subtitution
                
        cache_dp = dp
        return cache_dp
    
    # Nếu cache_dp đã được khởi tạo
    if len(cache_dp[0]) <= n:
        old_n = len(cache_dp[0]) - 1
        new_dp = [row[:] for row in cache_dp]
        for i in range(m + 1):
            for j in range(old_n + 1, n + 1):
                new_dp[i].append(0)
                if i == 0:
                    new_dp[0][j] = j
                else:
                    val = 2 if intersect_characters(str1[i-1], str2[j-1]) == 0 else 0
                    new_dp[i][j] = min( new_dp[i-1][j] + 1,
                                        new_dp[i][j-1] + 1,
                                        new_dp[i-1][j-1] + val)
        
        cache_dp = new_dp
        return new_dp
    
    return cache_dp

# Xóa cache của ma trận Levenshtein
def clear_cache_dp():
    global cache_dp
    cache_dp = None

# Các biến cache cho việc lưu trữ kết quả của các hàm tìm kiếm dữ liệu
similar_cache = {}
sino_cache = {}
intersect_cache = {}

# Tìm các từ đồng nghĩa của chữ hán nôm
def find_similar(sinonom_word, data):
    global similar_cache
    if sinonom_word in similar_cache:
        return similar_cache[sinonom_word]
    
    for entry in data:
        if entry['input'] == sinonom_word:
            similar_words = entry['similar']
            if isinstance(similar_words, str):
                similar_words = ast.literal_eval(similar_words)
                similar_words.insert(0, sinonom_word)
            similar_cache[sinonom_word] = set(similar_words)
            return similar_cache[sinonom_word]
        
    similar_cache[sinonom_word] = set()
    return similar_cache[sinonom_word]

# Tìm các chữ hán nôm tương ứng với chữ quốc ngữ
def find_sino(quocngu_word, data):
    global sino_cache
    if quocngu_word in sino_cache:
        return sino_cache[quocngu_word]
    
    sino_cache[quocngu_word] = {entry['sinonom'] for entry in data if entry['quocngu'] == quocngu_word}
    return sino_cache[quocngu_word]

# Tìm số lượng ký tự giao nhau giữa 2 từ
def intersect_characters(sinonom_word, quocngu_word):
    global intersect_cache
    key = (sinonom_word, quocngu_word)
    if key in intersect_cache:
        return intersect_cache[key]
    
    S1 = find_similar(sinonom_word, sino_similar_dic)
    S2 = find_sino(quocngu_word, qn_sino_dic)
    intersect_len = len(S1 & S2)
    result = intersect_len if S1 and S2 else 0
    intersect_cache[key] = result
    return result

# Xóa cache của các hàm tìm kiếm dữ liệu
def clear_caches():
    global similar_cache, sino_cache, intersect_cache
    similar_cache.clear()
    sino_cache.clear()
    intersect_cache.clear()

# Truy vết đường đi của MED levenshtein và tô màu
def backtrace_levenshtein(dp, str1, str2):
    i, j = len(str1), len(str2)
    red = workbook.add_format({'color': 'red'})
    blue = workbook.add_format({'color': 'blue'})
    black = workbook.add_format({'color': 'black'})
    format_pairs = []
    color = {}

    # truy vết đường đi
    while i > 0 and j > 0:
        current_char1 = str1[i - 1]
        current_char2 = str2[j - 1]
        intersection_len = intersect_characters(current_char1, current_char2)

        if dp[i][j] == dp[i - 1][j - 1] + (2 if intersection_len == 0 else 0):
            if intersection_len == 0:
                color[current_char1] = red
            elif intersection_len >= 1:
                color[current_char1] = black
            # else:
            #     color[current_char1] = black
            i -= 1
            j -= 1
        elif dp[i][j] == dp[i - 1][j] + 1:
            color[str1[i - 1]] = red
            i -= 1
        else:
            j -= 1

    while i > 0:
        color[str1[i - 1]] = red
        i -= 1

    # xây dựng format_pairs để tô màu
    for char in str1:
        char_color = color.get(char, red)
        format_pairs.extend((char_color, char))

    format_pairs.extend((black, ' '))
    return format_pairs

# Truy vết đường đi của MED levenshtein đánh dấu
def backtrace_lenven_other(dp, str1, str2):
    i, j = len(str1), len(str2)
    red = workbook.add_format({'color': 'red'})
    blue = workbook.add_format({'color': 'blue'})
    black = workbook.add_format({'color': 'black'})
    format_pairs = []
    color = {}
    results = []
    count_correct = 0
    while i > 0 and j > 0:
        current_char1 = str1[i - 1]
        current_char2 = str2[j - 1]
        intersection_len = intersect_characters(current_char1, current_char2)

        if dp[i][j] == dp[i - 1][j - 1] + (2 if intersection_len == 0 else 0):
            if intersection_len == 0:
                color[current_char1] = red
                results.append({"han_nom": current_char1, "correct": False, "quoc_ngu": None})
            elif intersection_len > 1:
                color[current_char1] = blue
                results.append({"han_nom": current_char1,"correct": True, "quoc_ngu": current_char2})
                count_correct += 1
            else:
                results.append({"han_nom": current_char1, "correct": True, "quoc_ngu": current_char2})
                color[current_char1] = black
                count_correct += 1
            i -= 1
            j -= 1
        elif dp[i][j] == dp[i - 1][j] + 1:
            results.append({"han_nom": current_char1, "correct": False, "quoc_ngu": None})
            color[str1[i - 1]] = red
            i -= 1
        else:
            j -= 1
    
    while i > 0:
        results.append({"han_nom": str1[i - 1], "correct": False, "quoc_ngu": None})
        color[str1[i - 1]] = red
        i -= 1

    # xây dựng format_pairs để tô màu
    for char in str1:
        char_color = color.get(char, red)
        format_pairs.extend((char_color, char))

    format_pairs.extend((black, ' '))
    results.reverse()

    return format_pairs, results, count_correct/len(str1)  

def compute_similarity_score(text1, text2, text3):

    set1 = set(text1)
    set2 = set(text3)
    intersection = set1.intersection(set2)
    intersect_score = len(intersection) / max(len(set1), len(set2))

    str = preprocess_quocngu_text(text2)
    # Tính ma trận Levenshtein
    dp = compute_levenshtein_matrix(text1, str)
    _, _, score = backtrace_lenven_other(dp, text1, str)

    # Xóa bộ nhớ tạm
    clear_cache_dp()

    # Điều chỉnh điểm số dựa trên tỷ lệ độ dài
    return (score + 1) * (1/abs(len(text1) - len(str) if len(text1) != len(str) else 1 )) * (1 + intersect_score)

# Dóng hàng giữa dữ liệu OCR và dữ liệu trích xuất
def alignments(ocr_datas, extract_datas):
    if not ocr_datas or not extract_datas:
        print("Dữ liệu đầu vào rỗng.")
        return

    results = []

    for data1, data2 in zip(ocr_datas, extract_datas):
        print(f"[+] Processing file: {data1['file_name']}")
        ocr_data = data1['data']
        extract_data = data2['data']

        i, j = 0, 0
        data = []
        while i < len(ocr_data) and j < len(extract_data):
            ocr_text = ocr_data[i]['text']
            current_score = compute_similarity_score(ocr_text, extract_data[j]['quocngu'], extract_data[j]['hannom'])

            # Duyệt qua extract_data để tìm điểm số cao nhất
            best_match_idx = j
            for tmp_j in range(j + 1, len(extract_data)):
                score = compute_similarity_score(ocr_text, extract_data[tmp_j]['quocngu'], extract_data[tmp_j]['hannom'])
                if score > current_score:
                    current_score = score
                    best_match_idx = tmp_j
                else:
                    break  # Dừng lại nếu điểm không cải thiện
            
            data.append({
                "sinonom_ocr": ocr_text,
                "sinonom_char": extract_data[best_match_idx]['hannom'],
                "quocngu": extract_data[best_match_idx]['quocngu'],
                "image_box": ocr_data[i]['position']
            })

            # Cập nhật chỉ số
            i += 1
            j = best_match_idx + 1

        while i < len(ocr_data):
            data.append({
                "sinonom_ocr": ocr_data[i]['text'],
                "sinonom_char": "",
                "quocngu": "",
                "image_box": ocr_data[i]['position']
            })
            i += 1

        results.append({
            "file_name": data1['file_name'],
            "data": data
        })

    return results

def process_data():
    han_nom_data = read_input_file_han_nom(INPUT_FILE_PATH_HAN_NOM_KANDIANGUJI)
    quoc_ngu_data = read_input_file_quoc_ngu(INPUT_FILE_PATH_QUOC_NGU)
    results = alignments(han_nom_data, quoc_ngu_data)
    row = 1
    print("[+] Writing data to excel file...")
    for result in results:
        cnt = 1
        for data in result['data']:
            ID = result['file_name'].replace("_page", ".").replace(".png", ".") + f"{cnt:03}"
            dp = compute_levenshtein_matrix(data['sinonom_ocr'], preprocess_quocngu_text(data['quocngu']))
            format_pairs = backtrace_levenshtein(dp, data['sinonom_ocr'], preprocess_quocngu_text(data['quocngu']))
            clear_cache_dp()
            worksheet.write(row, COL_IMAGE_NAME, result['file_name'])
            worksheet.write(row, COL_ID, ID)
            worksheet.write_rich_string(row, COL_SINONOM_OCR, *format_pairs)
            worksheet.write(row, COL_SINONOM_CHAR, data['sinonom_char'])
            worksheet.write(row, COL_CHU_QUOC_NGU, data['quocngu'])
            worksheet.write(row, COL_IMAGE_BOX, str(data['image_box']))
            row += 1
            cnt += 1

    workbook.close()
    print("[+] Done! Open the file result.xlsx to see the results.")

def main():
    global workbook, worksheet, qn_sino_dic, sino_similar_dic
    workbook, worksheet = create_workbook(OUTPUT_FILE_PATH)
    set_headers()
    qn_sino_dic = load_dictionary(QUOCNGU_SINONOM_DIC_FILE, 'quocngu', 'sinonom')
    sino_similar_dic = load_dictionary(SINONOM_SIMILAR_DIC_FILE, 'input', 'similar')
    process_data()

if __name__ == "__main__":
    main()